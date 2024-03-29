import csv
import os
from openpyxl import load_workbook
# from configparser import ConfigParser


"""
This class will handle the task of parsing excel file and writing csv file.
"""


class CsvFileCreator:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def parse_excel(self):
        try:
            # create workbook object
            wb = load_workbook(self.excel_path)
            rangeData = []
            for ws in wb:
                for rows in ws.iter_rows(ws.min_row+1, ws.max_row, ws.min_column, ws.max_column):
                    rowData = []
                    for row in rows:
                        rowData.append(row.value)
                    # print(rowData)
                    rangeData.append(rowData)
            # print(rangeData)
            return rangeData
        except Exception as ex:
            pass

    def write_csv(self, file_name, input_data, delimiter, header):
        try:
            file_path = os.path.dirname(self.excel_path)
            full_path = os.path.join(file_path, file_name)
            print(full_path)
            # open the file for writing
            with open(full_path, newline='', mode='w') as File:
                writer = csv.writer(File, delimiter=delimiter)
                # write the header first.
                if header:
                    header = header.upper()
                    writer.writerow([header])
                writer.writerows(input_data)
            return True
        except Exception as ex:
            print(ex)
            return False









